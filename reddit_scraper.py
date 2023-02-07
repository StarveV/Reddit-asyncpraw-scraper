import asyncio
import datetime
import openpyxl
import asyncpraw
import time 

no_days = 7
end_date = datetime.datetime.now()
start_date = end_date - datetime.timedelta(days=no_days)

print (start_date)

# Time out for one request
timeout_time = 30

# Additional trial after not finding new submission
max_additional_trial = 10

# Get the current date and time
now = datetime.datetime.now()
print (f'start scraping: {now}')

# Set the filename for the .xlsx file
filename = f"wsb_scrape_{now.strftime('%Y-%m-%d_%H-%M-%S')}.xlsx"

# Create a new workbook
workbook = openpyxl.Workbook()

# Set the reddit client key
client_id='your_client_id'
client_secret='your_secert'
user_agent='your_agent'

# Create a new sheet
sheet = workbook.active
sheet.title = "scrape"

# Add the column headers
sheet.cell(row=1, column=1).value = "date"
sheet.cell(row=1, column=2).value = "title"
sheet.cell(row=1, column=3).value = "score"
sheet.cell(row=1, column=4).value = "post body"
sheet.cell(row=1, column=5).value = "number of comments"

async def make_request(scraped_post_ids, start_date, end_date):
    # Create an async PRAW client
    reddit = asyncpraw.Reddit(client_id=client_id, client_secret=client_secret, user_agent=user_agent, timeoout = timeout_time)

    # Set a flag to track whether any new submissions have been found
    found_new_submission = False

    no_new_submissions_counter = 0

    # Retrieve a list of submissions in the subreddit
    subreddit = await reddit.subreddit("wallstreetbets")

    # Iterate over the submissions using the async for loop
    async for submission in subreddit.new(limit=None):
        # Convert the timestamp to a datetime object
        timestamp = datetime.datetime.fromtimestamp(submission.created_utc)
        
        # Check if the submission was made in the past `no_days` days from the start date
        if (start_date <= timestamp ):
            # Check if the submission has already been scraped
            if submission.id not in scraped_post_ids and timestamp <= end_date:
                # If the submission has not been scraped, add its ID to the list
                scraped_post_ids.append(submission.id)
                print (timestamp, submission.title )

                # Set the flag to True, to indicate that a new submission has been found
                found_new_submission = True
                # Add the data to the sheet
                sheet.cell(row=sheet.max_row + 1, column=1).value = timestamp.strftime("%Y-%m-%d %H:%M:%S")
                sheet.cell(row=sheet.max_row, column=2).value = submission.title
                sheet.cell(row=sheet.max_row, column=3).value = submission.score
                sheet.cell(row=sheet.max_row, column=4).value = submission.selftext
                sheet.cell(row=sheet.max_row, column=5).value = submission.num_comments
            else:
                # If the submission has already been scraped, do not scrape it again
                continue
        else:
            # If no new submissions are found, check the value of the flag
            if not found_new_submission:
                no_new_submissions_counter += 1
                
                if no_new_submissions_counter >= max_additional_trial:
                # If the maximum number of tries have been reached, break the loop
                    print (f'No new submission counter: {no_new_submissions_counter}')
                    break # If the flag is still False, exit the loop
            
    # Close the SSL transport
    # Close the client session
    await reddit.close()
    
async def main_task():
    # Create an empty list to store the ID of the scraped submissions
    scraped_post_ids = []
    # Call the `make_request` function
    await make_request(scraped_post_ids, start_date, end_date)
    # Save the workbook to the specified filename
    workbook.save(filename)

# Start the timer
start_time = time.perf_counter()

# Create an event loop
loop = asyncio.get_event_loop()
# Run the main_task function
loop.run_until_complete(main_task())

# Print a message to the console
print(f"Scrape complete. Results saved to {filename}.")

# Stop the timer
end_time = time.perf_counter()

# Calculate the elapsed time
elapsed_time = end_time - start_time

# Print the elapsed time
print(f'Elapsed time: {elapsed_time:0.4f} seconds')
